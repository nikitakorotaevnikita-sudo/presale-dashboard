import io
import openpyxl
from src.models import StatusEvent
from src.export import build_export


def ev(rid, status, month, service="A", hours=None, work=0.0, dur=0.0):
    return StatusEvent(
        request_id=rid, request="r", org="o", product="P", scale="S",
        service=service, initiator="I", team="T", business_unit="BU",
        status=status, prev_status="", date_start="", date_end="",
        month=month, duration_rd=dur, work_duration_rd=work, hours=hours, note="",
    )


def test_export_has_all_metric_sheets():
    events = [ev("1", "Инициализация", 1), ev("1", "Принято", 2, hours=10.0)]
    data = build_export(events, dimension="услуга")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) == {
        "Поступило", "Проработано", "Ср. трудоемкость",
        "Ср. длительность", "Ср. на контроле"}


def test_export_values_placed_by_month():
    events = [ev("1", "Инициализация", 3, service="A")]
    data = build_export(events, dimension="услуга")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Поступило"]
    header = [c.value for c in ws[1]]
    assert header[3] == 3  # колонка месяца 3 (B=1,C=2,D=3)
    row = [c.value for c in ws[2]]
    assert row[0] == "A"
    assert row[3] == 1
