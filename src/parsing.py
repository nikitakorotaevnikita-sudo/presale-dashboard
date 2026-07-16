import html as _html
import re
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter

from src.config import SHEET_NAME
from src.models import StatusEvent

REQUIRED_COLUMNS = [
    "Месяц начала текущего статуса",
    "Длительность проработки запроса (раб. дн.)",
    "Продукт", "Масштаб", "Услуга", "ИД запроса", "Запрос на пресейл",
    "Организация", "Инициатор", "Команда", "Бизнес-единица",
    "Предыдущий статус", "Статус", "Дата начала статуса",
    "Дата окончания статуса", "Длительность, р.д.",
    "Отработано часов", "Примечание",
]


class ParseError(Exception):
    pass


def _f(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _f0(v):
    """float со значением по умолчанию 0.0 (None трактуется как 0.0)."""
    f = _f(v)
    return f if f is not None else 0.0


def _i(v):
    f = _f(v)
    return int(f) if f is not None else None


def _s(v):
    return "" if v is None else str(v).strip()


def _rel_map(xmltext):
    """Id -> Target из *.rels (порядок атрибутов не важен)."""
    out = {}
    for rel in re.findall(r"<Relationship\b[^>]*>", xmltext):
        i = re.search(r'\bId="([^"]+)"', rel)
        t = re.search(r'\bTarget="([^"]+)"', rel)
        if i and t:
            out[i.group(1)] = t.group(1)
    return out


def _extract_hyperlinks(path, col_letter):
    """{номер строки Excel -> URL} по гиперссылкам ячеек колонки col_letter листа.

    openpyxl в read_only-режиме не отдаёт гиперссылки, поэтому читаем их напрямую
    из xlsx (zip): лист -> его *.rels -> внешние Target-ы.
    """
    try:
        z = zipfile.ZipFile(path)
    except (zipfile.BadZipFile, OSError):
        return {}
    try:
        wb = z.read("xl/workbook.xml").decode("utf-8", "ignore")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
        relmap = _rel_map(rels)
        sheet_rid = None
        for m in re.finditer(r"<sheet\b([^>]*)>", wb):
            attrs = m.group(1)
            nm = re.search(r'\bname="([^"]+)"', attrs)
            rid = re.search(r'r:id="([^"]+)"', attrs)
            if nm and rid and nm.group(1) == SHEET_NAME:
                sheet_rid = rid.group(1)
                break
        target = relmap.get(sheet_rid) if sheet_rid else None
        if not target:
            return {}
        sheet_path = target if target.startswith("xl/") else "xl/" + target.lstrip("/")
        try:
            xml = z.read(sheet_path).decode("utf-8", "ignore")
        except KeyError:
            return {}
        base = sheet_path.rsplit("/", 1)[-1]
        rels_path = sheet_path.rsplit("/", 1)[0] + "/_rels/" + base + ".rels"
        try:
            rid_to_url = _rel_map(z.read(rels_path).decode("utf-8", "ignore"))
        except KeyError:
            rid_to_url = {}
        out = {}
        for m in re.finditer(r"<hyperlink\b([^>]*?)/?>", xml):
            attrs = m.group(1)
            ref = re.search(r'\bref="([^"]+)"', attrs)
            rid = re.search(r'r:id="([^"]+)"', attrs)
            if not ref or not rid:
                continue
            cell = ref.group(1).split(":")[0]  # первая ячейка диапазона
            cm = re.match(r"([A-Z]+)(\d+)$", cell)
            if not cm or cm.group(1) != col_letter:
                continue
            url = rid_to_url.get(rid.group(1))
            if url:
                out[int(cm.group(2))] = _html.unescape(url)
        return out
    finally:
        z.close()


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ParseError(f"В файле нет листа «{SHEET_NAME}»")
        ws = wb[SHEET_NAME]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        idx = {name: i for i, name in enumerate(header) if name is not None}
        missing = [c for c in REQUIRED_COLUMNS if c not in idx]
        if missing:
            raise ParseError("В листе отсутствуют колонки: " + ", ".join(missing))

        def g(row, name):
            i = idx[name]
            return row[i] if i < len(row) else None

        # гиперссылки колонки «Запрос на пресейл» -> {строка Excel: URL}
        req_col = get_column_letter(idx["Запрос на пресейл"] + 1)
        links = _extract_hyperlinks(path, req_col)

        events = []
        # данные идут со 2-й строки Excel (после заголовка)
        for row_num, row in enumerate(it, start=2):
            rid = g(row, "ИД запроса")
            if rid in (None, ""):
                continue
            events.append(StatusEvent(
                request_id=_s(rid),
                request=_s(g(row, "Запрос на пресейл")),
                org=_s(g(row, "Организация")),
                product=_s(g(row, "Продукт")),
                scale=_s(g(row, "Масштаб")),
                service=_s(g(row, "Услуга")),
                initiator=_s(g(row, "Инициатор")),
                team=_s(g(row, "Команда")),
                business_unit=_s(g(row, "Бизнес-единица")),
                status=_s(g(row, "Статус")),
                prev_status=_s(g(row, "Предыдущий статус")),
                date_start=_s(g(row, "Дата начала статуса")),
                date_end=_s(g(row, "Дата окончания статуса")),
                month=_i(g(row, "Месяц начала текущего статуса")),
                duration_rd=_f0(g(row, "Длительность, р.д.")),
                work_duration_rd=_f0(g(row, "Длительность проработки запроса (раб. дн.)")),
                hours=_f(g(row, "Отработано часов")),
                note=_s(g(row, "Примечание")),
                link=links.get(row_num, ""),
            ))
        return events
    finally:
        wb.close()
